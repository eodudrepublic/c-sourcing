import os.path

from selenium.webdriver.common.by import By  # 웹 페이지의 요소를 지정하는 방법을 정의
import datetime as dt  # 날짜 및 시간 관련 추가 기능 제공
import time  # 시간과 관련된 기능 제공
import openpyxl  # 엑셀 파일을 읽고 쓰는데 사용
from urllib.parse import urlparse  # URL을 구문 분석하는데 사용
import re  # 정규 표현식을 사용하기 위한 모듈
from market_analysis.naver_store_analysis import competitor_sourcing
from sourcing.check_banned_words import check_words_in_sentence
from sourcing.convert_url import convert_url
from sourcing.get_category import soup_category
from sourcing.get_product_id import url_parser
from sourcing.taobao_img_search import taobao_searching


def taobao_sourcing(competitor, banned_words, _browser):
    current_time = dt.datetime.now()  # current_time : 시스템의 로컬 시간대 기준 현재 시간
    wb = openpyxl.Workbook()  # 새로운 엑셀 워크북 생성
    print("check---------------------------------")
    # sheet1은 활성화된 시트를 선택
    sheet1 = wb.active
    # sheet1의 시트이름 변경
    sheet1.title = "해외직구 베스트 소싱"
    # 엑셀 파일의 첫 번째 행(헤더)에는 각 열의 제목을 쓴다.
    sheet1.cell(row=1, column=1).value = '번호'
    sheet1.cell(row=1, column=2).value = '경쟁사 영어이름'
    sheet1.cell(row=1, column=3).value = '베스트 상품'
    sheet1.cell(row=1, column=4).value = '가격'
    sheet1.cell(row=1, column=5).value = '링크'
    sheet1.cell(row=1, column=6).value = '이미지 링크'
    sheet1.cell(row=1, column=7).value = '배송비'
    sheet1.cell(row=1, column=8).value = '타오바오 링크'
    sheet1.cell(row=1, column=9).value = '타오바오 제품번호'
    sheet1.cell(row=1, column=10).value = '네이버 카테고리'
    sheet1.cell(row=1, column=11).value = '타오바오 가격'
    # excel 리스트의 각 원소는 하나의 제품 정보를 문자열로 저장하고 있다.
    # 이 문자열을 분할하여 각 열에 적절하게 저장한다.

    # 주어진 키워드로 베스트 제품 링크 수집 -> competitor_sourcing 함수 호출
    best_list = competitor_sourcing(competitor, _browser)

    link_cnt = 0  # 링크 카운트 초기화
    for link in best_list:
        try:
            _browser.get(link)  # Selenium을 사용하여 해당 링크 열기
            _browser.implicitly_wait(10)  # 페이지 로딩 대기

            # 페이지 내 요소들 추출
            element_img = _browser.find_element(By.XPATH,
                                               '//*[@id="content"]/div/div[2]/div[1]/div[1]/div[2]/img').get_attribute(
                'src')
            element_page = _browser.find_element(By.CSS_SELECTOR, 'div._3k440DUKzy > div:nth-child(1) > h3').text

            # 상표권 침해 여부 확인
            if check_words_in_sentence(element_page, banned_words) != 0:
                print("침해상품입니다!")
                continue

            # 경쟁사 정보 및 제품 정보 추출 및 출력
            # ---------------경쟁사-------------------------------------------#
            parsed_link = urlparse(_browser.current_url)
            path = parsed_link.path
            if path.startswith('/'):
                link_segment = path.split('/')[1]
            else:
                link_segment = path.split('/')[0]
            # ----------------------------------------------------------------#
            # 정규 표현식을 사용하여 숫자만 추출
            product_number_match = re.search(r'/products/(\d+)', link)
            if product_number_match:
                product_number = product_number_match.group(1)
                print("추출된 상품 번호:", product_number)
            else:
                print("상품 번호를 찾을 수 없습니다.")

            element_price = _browser.find_element(By.XPATH,
                                                 '//*[@id="content"]/div/div[2]/div[2]/fieldset/div[1]/div[2]/div/strong/span[2]').text
            element_delivery = _browser.find_element(By.XPATH,
                                                    '//*[@id="content"]/div/div[2]/div[2]/fieldset/div[5]/div/span[2]').text
            if "무료배송" in element_delivery:
                element_delivery = '0'
            elif '(주문시 결제)' in element_delivery:
                element_delivery = re.sub(r'[^0-9]', '', element_delivery)
            time.sleep(1)
            category = soup_category(_browser)  # 카테고리 함수에서 return 받아옴
            # print(f"{link_segment}|{element_page}|{element_price}|{link}|{element_img}|{element_delivery}")
            taobao_url, taobao_price = taobao_searching(element_img, _browser)  # taobao_searching 함수 호출
            print("check point 0")
            taobao_url_re = convert_url(taobao_url)
            item_num = url_parser(taobao_url_re)

            print(
                f"{link_segment}|{element_page}|{element_price}|{link}|{element_img}|{element_delivery}|{taobao_url_re}|{item_num}|{category}|{taobao_price}")

            # 추출한 정보를 엑셀 시트에 작성
            sheet1.cell(row=link_cnt + 2, column=1).value = str(link_cnt + 1)  # 번호
            sheet1.cell(row=link_cnt + 2, column=2).value = str(link_segment)  # URL
            sheet1.cell(row=link_cnt + 2, column=3).value = element_page  # Image URL
            sheet1.cell(row=link_cnt + 2, column=4).value = element_price  # Title
            sheet1.cell(row=link_cnt + 2, column=5).value = link  # Price
            sheet1.cell(row=link_cnt + 2, column=6).value = element_img  # Delivery
            sheet1.cell(row=link_cnt + 2, column=7).value = element_delivery  # Delivery
            sheet1.cell(row=link_cnt + 2, column=8).value = taobao_url_re  # Delivery
            sheet1.cell(row=link_cnt + 2, column=9).value = item_num  # Delivery
            sheet1.cell(row=link_cnt + 2, column=10).value = category  # Delivery
            sheet1.cell(row=link_cnt + 2, column=11).value = taobao_price
            link_cnt += 1  # 링크 카운트 증가

        except:
            print(f"error service:")
            pass

        # 배열 요소 너무 많음

    # 엑셀 파일 저장 경로 설정 -> '프로젝트 폴더 경로/sourcing_data/'에 저장됨
    save_path = './sourcing_data/'

    # 저장 경로가 존재하지 않으면 생성
    if not os.path.exists(save_path):
        os.mkdir(save_path)

    wb.save(f"{save_path}{current_time.year}.{current_time.month}.{current_time.day}_{competitor.split('/')[3]}.xlsx")  # 엑셀 파일 저장
    print("Excel 파일 저장 완료")

    return 0
